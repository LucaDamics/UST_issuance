"""Parse CBO's detailed Feb 2026 baseline projections into seasonality buckets.

Inputs (data/):
  51142-2026-02-Spending-Projections.xlsx  - account-level (TAFS) outlay projections
  51138-2026-02-Revenue.xlsx               - revenue projections by source

Each Treasury account row is assigned to one of the model's seasonality buckets
using CBO's own "Major spending category" plus the budget function code. The
buckets are chosen to line up with groupings of the DTS mapping's cbo_category
column, so DTS-derived monthly shares can be applied to CBO annual totals.

Writes cbo_projection_buckets.csv: bucket x fiscal_year -> USD mn.
"""

import openpyxl
import pandas as pd

FYS = list(range(2026, 2037))
OUTLAY_COLS = {fy: 22 + i for i, fy in enumerate(FYS)}  # outlays block


def spending_bucket(category, function, disc):
    if category == "Social Security":
        return "social_security"
    if category == "Medicare":
        return "medicare_net"
    if category == "Medicaid":
        return "medicaid"
    if category == "Net Interest":
        return "net_interest"
    if category == "Defense":
        return "defence"
    if category == "Offsetting receipts":
        # Medicare premium offsets belong with Medicare so both sides are net
        return "medicare_net" if function == "570" else "nondefense_other"
    # Nondefense discretionary and Other (mandatory) spending: split by function
    if function == "700":
        return "veterans"
    if function == "600":
        return "income_security_retirement"
    if function == "500":
        return "education_training"
    if function == "550":
        return "health_other"
    return "nondefense_other"


def parse_spending():
    wb = openpyxl.load_workbook("data/51142-2026-02-Spending-Projections.xlsx")
    ws = wb["1. February 2026 Baseline"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for r in rows[8:]:
        if r[0] is None or r[3] is None:
            continue
        bucket = spending_bucket(str(r[3]), str(r[6]), str(r[2]))
        for fy, col in OUTLAY_COLS.items():
            v = r[col]
            if v is None:
                continue
            out.append({"side": "outlay", "bucket": bucket, "fy": fy, "usd_mn": float(v)})
    return pd.DataFrame(out).groupby(["side", "bucket", "fy"], as_index=False)["usd_mn"].sum()


REVENUE_ROWS = {
    "Individual income taxes": "individual_income",
    "Payroll taxes": "payroll",
    "Corporate income taxes": "corporate",
    "Customs duties": "customs",
    "Excise taxes": "excise",
    "Federal Reserve remittances": "misc_receipts",
    "Estate and gift taxes": "estate_gift",
    "Miscellaneous fees and fines": "misc_receipts",
}


def parse_revenue():
    wb = openpyxl.load_workbook("data/51138-2026-02-Revenue.xlsx")
    ws = wb["1. Revenue Projections"]
    rows = list(ws.iter_rows(values_only=True))
    out = []
    seen = set()  # the sheet repeats the same labels in a %-of-GDP block below
    for r in rows:
        name = str(r[0]).strip() if r[0] else ""
        if name in REVENUE_ROWS and name not in seen:
            seen.add(name)
            for i, fy in enumerate(FYS):
                v = r[2 + i]  # col1 = actual 2025; cols 2.. = 2026..2036
                if v is None:
                    continue
                out.append({"side": "receipt", "bucket": REVENUE_ROWS[name],
                            "fy": fy, "usd_mn": float(v) * 1000})  # $bn -> $mn
    df = pd.DataFrame(out).groupby(["side", "bucket", "fy"], as_index=False)["usd_mn"].sum()
    # the DTS cannot split withheld income tax from FICA, so the seasonal
    # profile is shared: project the combined line, keep components for levels
    combined = df[df.bucket.isin(["individual_income", "payroll"])] \
        .groupby("fy", as_index=False)["usd_mn"].sum()
    combined["side"] = "receipt"
    combined["bucket"] = "income_payroll"
    return pd.concat([df, combined], ignore_index=True)


def main():
    sp = parse_spending()
    rv = parse_revenue()
    df = pd.concat([sp, rv], ignore_index=True)
    df.to_csv("cbo_projection_buckets.csv", index=False)

    p = df.pivot_table(index=["side", "bucket"], columns="fy", values="usd_mn", aggfunc="sum")
    pd.set_option("display.float_format", lambda x: f"{x/1000:,.0f}")
    print("CBO Feb 2026 baseline by seasonality bucket (USD bn):")
    print(p[[2026, 2027, 2030, 2036]].to_string())
    tot_out = p.loc["outlay"].sum()
    tot_rec = p.loc[("receipt",)].drop(index=["individual_income", "payroll"]).sum()
    print("\nCheck FY2026: outlays %.0f bn, receipts %.0f bn, deficit %.0f bn"
          % (tot_out[2026] / 1e3, tot_rec[2026] / 1e3, (tot_out[2026] - tot_rec[2026]) / 1e3))


if __name__ == "__main__":
    main()
